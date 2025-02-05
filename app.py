import os

from langchain_groq import ChatGroq
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import JsonOutputParser
import json

import streamlit as st
from pprint import pprint

import win32com.client
import openpyxl

import pythoncom
pythoncom.CoInitialize()


st.set_page_config(page_title="Invoice-Exctarctor")
st.header("Invoices Extractor")

from dotenv import load_dotenv
load_dotenv()

os.environ["GROQ_API_KEY"] = os.getenv("GROQ_API_KEY")

# Initialize Groq LLM
llm = ChatGroq(
    model_name="mixtral-8x7b-32768",
    temperature=0.5
)

# Define the expected JSON structure
parser = JsonOutputParser(pydantic_object={
    "type": "object",
    "properties": {
        
    }
})

prompt = ChatPromptTemplate.from_messages([
    ("system", """You are a helpful assistant which helps to extract the order details from email and turn it into a JSON file. Extract invoice details such as name, address, city, order value, etc., and give it in a JSON file. Keep the output structure as this only.

    <output_keys>
        requestor_details.name
        requestor_details.department
        requestor_details.reason
        requestor_details.domain
        identifying_information.supplier_name
        identifying_information.gstin_number
        identifying_information.gst_rate
        identifying_information.address.street_address
        identifying_information.address.city
        identifying_information.address.state
        identifying_information.address.postal_code
        identifying_information.address.country_code
        identifying_information.contact.name
        identifying_information.contact.email_id
        identifying_information.contact.email_id_finance
        identifying_information.contact.phone_number
        identifying_information.pan_no
        identifying_information.msme_registration_no
        identifying_information.hsn_sac_code
        identifying_information.supplier_type
        local_vendor_account_information.beneficiary_account_name
        local_vendor_account_information.payment_currency
        local_vendor_account_information.bank_name
        local_vendor_account_information.ifsc_code
        local_vendor_account_information.account_type
        local_vendor_account_information.beneficiary_account_number
        local_vendor_account_information.bank_routing_method
        local_vendor_account_information.remittance_email
        foreign_currency_vendor_account_information.beneficiary_account_name
        foreign_currency_vendor_account_information.payment_currency
        foreign_currency_vendor_account_information.bank_name
        foreign_currency_vendor_account_information.account_type
        foreign_currency_vendor_account_information.swift_code
        foreign_currency_vendor_account_information.iban_no
        foreign_currency_vendor_account_information.bank_routing_method
        foreign_currency_vendor_account_information.intermediary_bank_routing_method
        foreign_currency_vendor_account_information.intermediary_swift_no
        foreign_currency_vendor_account_information.intermediary_bank_name
        foreign_currency_vendor_account_information.remittance_email
    </output_keys>
    """),
    ("user", "{input}")
])


# Create the chain that guarantees JSON output
chain = prompt | llm | parser

def parse_mail(mail: str) -> dict:
    result = chain.invoke({"input": mail})
    return json.dumps(result, indent=2)

mail = st.text_input("Enter Email/Text Invoice")

submit = st.button("Extract")

if submit:
    invoice = parse_mail(mail)
    print(invoice)
    # st.json(invoice, expanded=True)
    st.json(invoice)
    invoice = json.loads(invoice)

    with open('cell_no.json') as cell:
        cell = json.load(cell)
    print(type(cell))

    wb = openpyxl.load_workbook("excel_file/Supplier_Form.xlsx")
    ws = wb.active # First Sheet


    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

    print('The value in cell B10 is: ', ws['B10'].value)

    print("checkpoint 3")

    for val in invoice:
        for j in invoice[val]:
            if isinstance(invoice[val][j], dict):
                for k in invoice[val][j]:
                    cell_no = cell[val][j][k]
                    value = invoice[val][j][k]
                    print(k, " ", value, cell_no)
                    if ws[cell_no].value is None:
                        ws[cell_no].value = value
                    else:
                        ws[cell_no].value += value
            else:
                cell_no = cell[val][j]
                value = invoice[val][j]
                print(j, " ", value, cell_no)
                if ws[cell_no].value is None:
                    ws[cell_no].value = value
                else:
                    ws[cell_no].value += value

    print("checkpoint 4")
    wb.save("Supplier.xlsx")
    wb.close()

    print("checkpoint 5")
    pythoncom.CoInitialize()
    # Open Excel Application
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False  # Run in background

    # Open Workbook
    wb = excel.Workbooks.Open(r"D:Playground/Invoices-Extraction/Supplier.xlsx")

    # Export as PDF (change path as needed)
    pdf_name = f"invoice.pdf"
    pdf_path = f"D:/Playground/Invoices-Extraction/PDF/{pdf_name}"
    wb.ExportAsFixedFormat(0, pdf_path)

    # Close Workbook and Quit Excel
    wb.Close(False)
    excel.Quit()
    # Uninitialize COM
    pythoncom.CoUninitialize()
    print(f"Excel file converted to PDF: {pdf_path}")
    
    # pdf_path = os.path.join(pdf_folder, pdf_name)

    if os.path.exists(pdf_path):
        with open(pdf_path, "rb") as pdf_file:
            pdf_bytes = pdf_file.read()

        st.success("✅ PDF Generated Successfully!")
        
        # Download button
        st.download_button(
            label="📥 Download PDF",
            data=pdf_bytes,
            file_name=pdf_name,
            mime="application/pdf"
        )
    else:
        st.error("⚠️ PDF file not found. Please generate it first.")
