import json
import openpyxl

with open('invoice.json') as invoicejson:
    invoice = json.load(invoicejson)

with open('cell_no.json') as cell:
    cell = json.load(cell)

def json_to_pdf(invoice, cell):
    wb = openpyxl.load_workbook("excel_file/Supplier_Form.xlsx")
    ws = wb.active # First Sheet


    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

    print('The value in cell B10 is: ', ws['B10'].value)


    for val in invoice:
        for j in invoice[val]:
            if isinstance(invoice[val][j], dict):
                for k in invoice[val][j]:
                    cell_no = cell[val][j][k]
                    value = invoice[val][j][k]
                    print(k, " ", value, cell_no)
                    if ws[cell_no].value is None:
                        ws[cell_no].value = value
                    else:
                        ws[cell_no].value += value
            else:
                cell_no = cell[val][j]
                value = invoice[val][j]
                print(j, " ", value, cell_no)
                if ws[cell_no].value is None:
                    ws[cell_no].value = value
                else:
                    ws[cell_no].value += value

    wb.save("Supplier.xlsx")
    wb.close()

    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client

    # Open Excel Application
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False  # Run in background

    # Open Workbook
    wb = excel.Workbooks.Open(r"D:/Playground/Invoices-Extraction/Supplier.xlsx")

    # Export as PDF (change path as needed)
    pdf_path = r"D:/Playground/Invoices-Extraction/PDF/Supplier.pdf"
    wb.ExportAsFixedFormat(0, pdf_path)

    # Close Workbook and Quit Excel
    wb.Close()
    excel.Quit()

    # Uninitialize COM
    pythoncom.CoUninitialize()

    print(f"Excel file converted to PDF: {pdf_path}")


json_to_pdf(invoice, cell)